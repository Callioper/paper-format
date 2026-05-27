#!/usr/bin/env python3
"""verify_quotes.py - Direct quotation verification against source PDFs.

Extracts direct quotations from a thesis .docx, matches them against
source PDF documents, and produces a verification report with evidence.

Supports text-based PDFs (direct text extraction). Scanned PDF OCR
support planned for v3.1.

Usage:
    python verify_quotes.py thesis.docx --sources "ref1.pdf,ref2.pdf" --output quote_result.json
    python verify_quotes.py thesis.docx --sources "book.pdf:1-50" --output quote_result.json
    python verify_quotes.py thesis.docx --extract-only  # preview extracted quotes
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

_THIS = Path(__file__).resolve()
_REPO_ROOT = _THIS.parents[1]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from scripts.lib.quote_models import QuoteRecord, QuoteVerificationResult


# -------------------------------------------------------------------
# Quotation extraction from thesis .docx
# -------------------------------------------------------------------

# Patterns for direct quotations in Chinese academic writing
QUOTE_PATTERNS = [
    re.compile(r'“([^”]{4,})”'),          # "xxx" (Unicode curly)
    re.compile(r'‘([^’]{4,})’'),          # 'xxx' (Unicode single curly)
    re.compile(r'「([^」]{4,})」'),          # 「xxx」 (CJK corner brackets)
    re.compile(r'『([^』]{4,})』'),          # 『xxx』 (double CJK corner)
    re.compile(r'"([^"]{4,})"'),                          # "xxx" (ASCII quotes, Chinese context)
]

# Section headers that mark the end of body text
STOP_SECTIONS = re.compile(
    r'^(参考文献|References|Bibliography|致\s*谢|附\s*录|Appendix)', re.IGNORECASE
)


def extract_quotes_from_docx(docx_path: str | Path) -> list[QuoteRecord]:
    """Extract all direct quotations from a thesis .docx.

    Returns list of QuoteRecord with quote text, context, and position.
    """
    from docx import Document

    doc = Document(docx_path)
    quotes: list[QuoteRecord] = []
    idx = 0

    for para_num, para in enumerate(doc.paragraphs):
        text = para.text.strip()
        if not text:
            continue

        # Stop at references section
        if STOP_SECTIONS.match(text):
            break

        # Skip heading paragraphs
        style_name = para.style.name if para.style else ""
        if style_name.startswith("Heading") or style_name.startswith("TOC"):
            continue

        # Find all quotations in this paragraph
        for pattern in QUOTE_PATTERNS:
            for match in pattern.finditer(text):
                quote_text = match.group(1).strip()

                # Skip very short matches (likely not real quotations)
                if len(quote_text) < 4:
                    continue

                # Skip if it looks like a reference entry
                if re.match(r'^\d+', quote_text):
                    continue

                idx += 1
                start, end = match.start(), match.end()
                context_before = text[max(0, start - 30):start]
                context_after = text[end:min(len(text), end + 30)]

                # Try to extract source hint from footnote markers nearby
                source_hint = _extract_source_hint(text, end)

                quotes.append(QuoteRecord(
                    index=idx,
                    quote_text=quote_text,
                    context_before=context_before,
                    context_after=context_after,
                    paragraph_index=para_num,
                    source_hint=source_hint,
                ))

    return quotes


def _extract_source_hint(text: str, quote_end: int) -> str:
    """Try to extract author/year/page hint from text after the quote.

    Looks for patterns like: ——福柯《规训与惩罚》第30页
    or footnote markers like ① [1]
    """
    after = text[quote_end:quote_end + 60]

    # Pattern: ——Author, Title, p. X  or  ——作者《书名》第X页
    hint_match = re.search(r'[——\-—]\s*(.{2,30}?)(?:第\s*(\d+)\s*页)?', after)
    if hint_match:
        return hint_match.group(0).strip()

    # Pattern: footnote marker ①②③ or [1]
    fn_match = re.search(r'[①②③④⑤⑥⑦⑧⑨⑩]|\[\d+\]', after)
    if fn_match:
        return f"脚注{fn_match.group(0)}"

    return ""


# -------------------------------------------------------------------
# PDF text extraction
# -------------------------------------------------------------------

def parse_source_spec(source_str: str) -> tuple[Path, tuple[int, int] | None]:
    """Parse source specification like 'book.pdf:1-50'.

    Returns (path, (start_page, end_page) or None).
    """
    parts = source_str.split(":")
    pdf_path = Path(parts[0].strip())

    page_range = None
    if len(parts) > 1:
        range_str = parts[1].strip()
        range_match = re.match(r'(\d+)\s*-\s*(\d+)', range_str)
        if range_match:
            page_range = (int(range_match.group(1)), int(range_match.group(2)))
        elif range_str.isdigit():
            page_range = (int(range_str), int(range_str))

    return pdf_path, page_range


def extract_pdf_text(
    pdf_path: str | Path,
    page_range: tuple[int, int] | None = None,
    use_ocr: bool = False,
) -> dict[int, str]:
    """Extract text from each page of a PDF file.

    Args:
        pdf_path: Path to PDF file
        page_range: Optional (start, end) page range (1-indexed)
        use_ocr: If True, use PaddleOCR for scanned PDFs instead of direct text extraction

    Returns dict mapping page_number (1-indexed) -> page text.
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists() or pdf_path.suffix.lower() != ".pdf":
        return {}

    if use_ocr:
        return _extract_pdf_text_ocr(pdf_path, page_range)

    return _extract_pdf_text_plumber(pdf_path, page_range)


def _extract_pdf_text_plumber(pdf_path: Path, page_range: tuple[int, int] | None) -> dict[int, str]:
    """Direct text extraction via pdfplumber (for text-based PDFs)."""
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError("pdfplumber is required. Install with: pip install pdfplumber")

    page_texts: dict[int, str] = {}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            start = (page_range[0] - 1) if page_range else 0
            end = page_range[1] if page_range else len(pdf.pages)

            for i in range(start, min(end, len(pdf.pages))):
                page = pdf.pages[i]
                text = page.extract_text() or ""
                if text.strip():
                    page_texts[i + 1] = text
    except Exception as e:
        print(f"[verify_quotes] Error reading {pdf_path}: {e}")

    return page_texts


def _extract_pdf_text_ocr(pdf_path: Path, page_range: tuple[int, int] | None) -> dict[int, str]:
    """OCR-based text extraction for scanned PDFs using PaddleOCR.

    Renders each page to image, then runs PaddleOCR.
    Falls back to pdfplumber direct extraction if PaddleOCR unavailable.
    """
    try:
        import pdfplumber
    except ImportError:
        return {}

    try:
        from paddleocr import PaddleOCR
    except ImportError:
        print("[verify_quotes] PaddleOCR not installed, falling back to direct extraction")
        return _extract_pdf_text_plumber(pdf_path, page_range)

    # Initialize PaddleOCR (Chinese + English)
    print("[verify_quotes] Initializing PaddleOCR (first run downloads model)...")
    ocr = PaddleOCR(use_angle_cls=True, lang="ch", show_log=False)

    page_texts: dict[int, str] = {}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            start = (page_range[0] - 1) if page_range else 0
            end = page_range[1] if page_range else len(pdf.pages)

            for i in range(start, min(end, len(pdf.pages))):
                page = pdf.pages[i]

                # First try direct extraction
                direct_text = (page.extract_text() or "").strip()
                if len(direct_text) > 50:
                    page_texts[i + 1] = direct_text
                    continue

                # Page has little/no text — render to image and OCR
                img = page.to_image(resolution=200)
                import tempfile, os
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    tmp_path = tmp.name
                    img.save(tmp_path)

                try:
                    result = ocr.ocr(tmp_path, cls=True)
                    lines = []
                    if result and result[0]:
                        for line in result[0]:
                            text = line[1][0] if isinstance(line[1], (list, tuple)) else str(line[1])
                            lines.append(text)
                    if lines:
                        page_texts[i + 1] = "\n".join(lines)
                finally:
                    os.unlink(tmp_path)

    except Exception as e:
        print(f"[verify_quotes] OCR error for {pdf_path}: {e}")

    return page_texts


def detect_scanned_pdf(pdf_path: str | Path, sample_pages: int = 3) -> bool:
    """Detect if a PDF is scanned (image-based) by checking text content.

    Returns True if the PDF appears to be scanned (very little extractable text).
    """
    try:
        import pdfplumber
    except ImportError:
        return False

    pdf_path = Path(pdf_path)
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_chars = 0
            pages_checked = min(sample_pages, len(pdf.pages))
            for i in range(pages_checked):
                text = pdf.pages[i].extract_text() or ""
                total_chars += len(text.strip())
            # If average less than 30 chars per page, likely scanned
            return (total_chars / max(pages_checked, 1)) < 30
    except Exception:
        return False


def capture_page_screenshot(
    pdf_path: str | Path,
    page_num: int,
    output_path: str | Path,
    highlight_text: str = "",
) -> bool:
    """Capture a specific page from a PDF as a PNG image.

    If highlight_text is provided, draws a semi-transparent yellow rectangle
    around the matching text region on the page.

    Args:
        pdf_path: Path to source PDF
        page_num: Page number (1-indexed)
        output_path: Where to save the PNG
        highlight_text: Text to highlight (optional)

    Returns True if successful.
    """
    try:
        import pdfplumber
    except ImportError:
        return False

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if page_num < 1 or page_num > len(pdf.pages):
                return False
            page = pdf.pages[page_num - 1]
            img = page.to_image(resolution=150)
            img.save(str(output_path))

            # Add highlight if text provided
            if highlight_text:
                _add_highlight_to_screenshot(output_path, page, highlight_text)

            return True
    except Exception:
        return False


def _add_highlight_to_screenshot(image_path: str | Path, pdfplumber_page, search_text: str) -> None:
    """Draw a semi-transparent yellow highlight rectangle on the screenshot
    where the search text appears on the PDF page.

    Uses pdfplumber's word extraction to find bounding boxes,
    then draws rectangles using Pillow.
    """
    try:
        from PIL import Image, ImageDraw, ImageFilter
    except ImportError:
        return

    # Find words on the page that match parts of the search text
    try:
        words = pdfplumber_page.extract_words()
    except Exception:
        return

    if not words:
        return

    # Normalize search text for matching
    search_clean = re.sub(r'\s+', '', search_text)

    # Find consecutive word sequences that match the search text
    highlight_bboxes = []
    for i in range(len(words)):
        # Try building a sequence starting from word i
        accumulated = ""
        for j in range(i, min(i + len(words), i + 20)):
            word_text = words[j].get("text", "")
            accumulated += re.sub(r'\s+', '', word_text)

            if len(accumulated) >= len(search_clean) * 0.7:
                # Check similarity
                from difflib import SequenceMatcher
                score = SequenceMatcher(None, search_clean[:len(accumulated)], accumulated).ratio()
                if score >= 0.85:
                    # Found a match — collect bounding box
                    x0 = min(w["x0"] for w in words[i:j+1])
                    y0 = min(w["top"] for w in words[i:j+1])
                    x1 = max(w["x1"] for w in words[i:j+1])
                    y1 = max(w["bottom"] for w in words[i:j+1])
                    highlight_bboxes.append((x0, y0, x1, y1))
                    break

    if not highlight_bboxes:
        return

    # Draw highlights on the image
    img = Image.open(image_path).convert("RGBA")
    overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    # Scale factor: pdfplumber uses points, image uses pixels at 150 DPI
    scale = 150 / 72.0  # 72 points per inch, 150 DPI

    for x0, y0, x1, y1 in highlight_bboxes:
        # Add padding
        pad = 3
        sx0 = int((x0 - pad) * scale)
        sy0 = int((y0 - pad) * scale)
        sx1 = int((x1 + pad) * scale)
        sy1 = int((y1 + pad) * scale)
        draw.rectangle([sx0, sy0, sx1, sy1], fill=(255, 255, 0, 80))  # Yellow, semi-transparent

    # Composite overlay onto original
    result = Image.alpha_composite(img, overlay)
    result.convert("RGB").save(str(image_path))


# -------------------------------------------------------------------
# Quote matching
# -------------------------------------------------------------------

def _normalize_for_match(text: str) -> str:
    """Normalize text for comparison: strip whitespace and punctuation variants."""
    text = re.sub(r'\s+', '', text)
    # Normalize quotation marks
    text = text.replace('“', '"').replace('”', '"')
    text = text.replace('「', '「').replace('」', '」')
    return text


def match_quote_against_source(
    quote: str,
    page_texts: dict[int, str],
    source_name: str,
    screenshots_dir: Path | None = None,
    pdf_path: Path | None = None,
) -> QuoteVerificationResult:
    """Match a single quotation against extracted PDF page texts.

    Uses exact substring match for long quotes, fuzzy match for shorter ones.
    """
    result = QuoteVerificationResult(
        quote_text=quote,
        source_file=source_name,
    )

    if not page_texts:
        result.status = "未定位到出处"
        result.note = "PDF 无可提取文本（可能是扫描版）"
        return result

    quote_norm = _normalize_for_match(quote)

    # Strategy 1: Exact substring match (for long quotes)
    if len(quote_norm) > 15:
        for page_num, page_text in page_texts.items():
            page_norm = _normalize_for_match(page_text)
            if quote_norm in page_norm:
                result.status = "一致"
                result.confidence = 1.0
                result.matched_page = page_num
                result.matched_text = quote

                # Capture screenshot with highlight
                if screenshots_dir and pdf_path:
                    screenshot_path = screenshots_dir / f"quote_{result.quote_index:03d}_p{page_num}.png"
                    if capture_page_screenshot(pdf_path, page_num, screenshot_path, highlight_text=quote):
                        result.screenshot_path = str(screenshot_path)
                return result

    # Strategy 2: Fuzzy match with sliding window
    best_score = 0.0
    best_page = 0
    best_context = ""

    for page_num, page_text in page_texts.items():
        page_norm = _normalize_for_match(page_text)

        # Skip very short pages (likely headers/footers)
        if len(page_norm) < 10:
            continue

        # Sliding window approach
        window_size = len(quote_norm) + 15
        step = max(1, len(quote_norm) // 3)

        for i in range(0, max(1, len(page_norm) - window_size + 1), step):
            window = page_norm[i:i + window_size]
            score = SequenceMatcher(None, quote_norm, window).ratio()

            if score > best_score:
                best_score = score
                best_page = page_num
                best_context = page_text[max(0, i // 2):min(len(page_text), i // 2 + len(quote) + 40)]

    # Determine status based on confidence
    if best_score >= 0.92:
        result.status = "一致"
    elif best_score >= 0.78:
        result.status = "基本一致"
    elif best_score >= 0.55:
        result.status = "疑似不一致"
    else:
        result.status = "未定位到出处"

    result.confidence = best_score
    result.matched_page = best_page
    result.matched_text = best_context[:200] if best_context else ""

    # Capture screenshot for matched page (with highlight)
    if best_page > 0 and screenshots_dir and pdf_path:
        screenshot_path = screenshots_dir / f"quote_{result.quote_index:03d}_p{best_page}.png"
        if capture_page_screenshot(pdf_path, best_page, screenshot_path, highlight_text=quote):
            result.screenshot_path = str(screenshot_path)

    if result.status == "基本一致":
        result.note = f"相似度 {best_score:.0%}，可能存在细微差异"
    elif result.status == "疑似不一致":
        result.note = f"相似度 {best_score:.0%}，建议人工核实"

    return result


# -------------------------------------------------------------------
# Main verification pipeline
# -------------------------------------------------------------------

def verify_quotes(
    thesis_path: str | Path,
    source_specs: list[str],
    output_dir: str | Path | None = None,
    force_ocr: bool = False,
) -> tuple[list[QuoteRecord], list[QuoteVerificationResult]]:
    """Run the full quote verification pipeline.

    Args:
        thesis_path: Path to thesis .docx
        source_specs: List of source specifications like "book.pdf" or "book.pdf:1-50"
        output_dir: Directory for screenshots and intermediate files
        force_ocr: Force OCR mode for all PDFs (auto-detected by default)

    Returns:
        (quotes, results) tuple
    """
    thesis_path = Path(thesis_path)
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        screenshots_dir = output_dir / "screenshots"
        screenshots_dir.mkdir(exist_ok=True)
    else:
        screenshots_dir = None

    # Step 1: Extract quotations from thesis
    print(f"[verify_quotes] Extracting quotations from {thesis_path.name}...")
    quotes = extract_quotes_from_docx(thesis_path)
    print(f"[verify_quotes] Found {len(quotes)} direct quotations")

    if not quotes:
        return quotes, []

    # Step 2: Parse source PDFs
    source_pages: dict[str, dict[int, str]] = {}
    source_paths: dict[str, Path] = {}

    for spec in source_specs:
        pdf_path, page_range = parse_source_spec(spec)
        if not pdf_path.exists():
            print(f"[verify_quotes] Warning: Source not found: {pdf_path}")
            continue

        # Auto-detect scanned PDFs
        use_ocr = force_ocr
        if not use_ocr:
            use_ocr = detect_scanned_pdf(pdf_path)
            if use_ocr:
                print(f"[verify_quotes] {pdf_path.name} appears to be scanned, using OCR")

        print(f"[verify_quotes] Extracting text from {pdf_path.name}{'(OCR)' if use_ocr else ''}...")
        pages = extract_pdf_text(pdf_path, page_range, use_ocr=use_ocr)
        source_name = pdf_path.stem
        source_pages[source_name] = pages
        source_paths[source_name] = pdf_path
        print(f"[verify_quotes]   {len(pages)} pages with text")

    if not source_pages:
        print("[verify_quotes] No source PDFs could be parsed")
        return quotes, []

    # Step 3: Match each quote against all sources
    print(f"[verify_quotes] Matching {len(quotes)} quotations against {len(source_pages)} sources...")
    results: list[QuoteVerificationResult] = []

    for quote in quotes:
        best_result = QuoteVerificationResult(
            quote_index=quote.index,
            quote_text=quote.quote_text,
            status="未定位到出处",
        )

        for source_name, pages in source_pages.items():
            result = match_quote_against_source(
                quote.quote_text,
                pages,
                source_name,
                screenshots_dir,
                source_paths.get(source_name),
            )
            result.quote_index = quote.index

            # Keep the best match across all sources
            if result.confidence > best_result.confidence:
                best_result = result

            # Perfect match — stop searching other sources
            if result.status == "一致":
                break

        results.append(best_result)

    # Summary
    status_counts = {}
    for r in results:
        status_counts[r.status] = status_counts.get(r.status, 0) + 1

    print(f"[verify_quotes] Results: {status_counts}")

    return quotes, results


# -------------------------------------------------------------------
# Excel output
# -------------------------------------------------------------------

def export_to_excel(
    results: list[QuoteVerificationResult],
    output_path: str | Path,
) -> None:
    """Export verification results to an Excel workbook with embedded screenshots.

    Columns: 序号, 论文中的引文, 出处文件, 出处页码, 核对结果, 出处截图, 备注
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")

    from PIL import Image as PILImage

    wb = Workbook()
    ws = wb.active
    ws.title = "引文核对"

    headers = ["序号", "论文中的引文", "出处文件", "出处页码", "核对结果", "出处截图", "备注"]
    ws.append(headers)

    # Header styling
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.fill = header_fill

    # Column widths
    widths = {"A": 8, "B": 44, "C": 20, "D": 10, "E": 14, "F": 30, "G": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Status color fills
    status_fills = {
        "一致": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "基本一致": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "疑似不一致": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "未定位到出处": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }

    for row_idx, r in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=r.quote_index)
        ws.cell(row=row_idx, column=2, value=r.quote_text[:120])
        ws.cell(row=row_idx, column=3, value=r.source_file)
        ws.cell(row=row_idx, column=4, value=r.matched_page if r.matched_page else "")
        ws.cell(row=row_idx, column=5, value=r.status)
        ws.cell(row=row_idx, column=7, value=r.note)

        # Apply status color
        status_cell = ws.cell(row=row_idx, column=5)
        fill = status_fills.get(r.status)
        if fill:
            status_cell.fill = fill

        # Wrap text for content columns
        for col_idx in (1, 2, 3, 4, 5, 7):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                vertical="top", wrap_text=True
            )

        # Embed screenshot if available
        if r.screenshot_path:
            ss_path = Path(r.screenshot_path)
            if ss_path.is_file():
                try:
                    from openpyxl.drawing.image import Image as XlImage
                    img = XlImage(str(ss_path))
                    # Scale to fit column width
                    max_width = 220
                    if img.width and img.width > max_width:
                        ratio = max_width / img.width
                        img.width = int(img.width * ratio)
                        img.height = int(img.height * ratio)
                    ws.add_image(img, f"F{row_idx}")
                    ws.row_dimensions[row_idx].height = max(110, (img.height or 100) * 0.75)
                except Exception:
                    ws.cell(row=row_idx, column=6, value=f"截图: {ss_path.name}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"[verify_quotes] Excel saved to {output_path}")


# -------------------------------------------------------------------
# CLI
# -------------------------------------------------------------------

def main():
    from scripts.lib import setup_utf8_output
    setup_utf8_output()

    parser = argparse.ArgumentParser(description="Direct quotation verification")
    parser.add_argument("thesis", help="Path to thesis .docx")
    parser.add_argument(
        "--sources", "-s",
        help="Comma-separated source PDF paths (optionally with page range: file.pdf:1-50)",
    )
    parser.add_argument("--output", "-o", help="Output JSON path")
    parser.add_argument(
        "--output-dir",
        help="Directory for screenshots and intermediate files",
    )
    parser.add_argument(
        "--extract-only",
        action="store_true",
        help="Only extract quotations, don't verify against sources",
    )
    parser.add_argument(
        "--ocr",
        action="store_true",
        help="Force OCR mode for scanned PDFs (auto-detected by default)",
    )
    parser.add_argument(
        "--excel",
        help="Export results to Excel workbook (.xlsx) with embedded screenshots",
    )
    args = parser.parse_args()

    if args.extract_only:
        quotes = extract_quotes_from_docx(args.thesis)
        print(f"\nExtracted {len(quotes)} quotations:")
        for q in quotes:
            print(f"  [{q.index}] \"{q.quote_text[:60]}\"")
            if q.source_hint:
                print(f"       hint: {q.source_hint}")
        if args.output:
            Path(args.output).write_text(
                json.dumps([q.to_dict() for q in quotes], ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            print(f"\nQuotes written to {args.output}")
        return

    if not args.sources:
        print("Error: --sources is required (or use --extract-only)")
        sys.exit(1)

    source_specs = [s.strip() for s in args.sources.split(",") if s.strip()]
    output_dir = args.output_dir or str(Path(args.thesis).parent / "quote_verification")

    quotes, results = verify_quotes(
        args.thesis, source_specs, output_dir, force_ocr=args.ocr
    )

    output = {
        "source_file": str(args.thesis),
        "source_pdfs": source_specs,
        "total_quotes": len(quotes),
        "total_verified": len(results),
        "status_summary": {
            status: sum(1 for r in results if r.status == status)
            for status in ("一致", "基本一致", "疑似不一致", "未定位到出处")
        },
        "results": [r.to_dict() for r in results],
    }

    if args.output:
        Path(args.output).write_text(
            json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"\nResults written to {args.output}")
    else:
        print(json.dumps(output, ensure_ascii=False, indent=2))

    # Excel export
    if args.excel and results:
        export_to_excel(results, args.excel)


if __name__ == "__main__":
    main()
